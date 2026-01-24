import logging
from typing import List, Dict

from PySide6.QtCore import QAbstractTableModel, QModelIndex, Slot, Qt
from PySide6.QtGui import QAction
from PySide6.QtWidgets import QTableView, QScrollBar, QMenu, QMessageBox, QFileDialog, QApplication
from openpyxl.styles import Font, Alignment
from openpyxl.workbook import Workbook

from app.user_data import DoIPMessageStruct

logger = logging.getLogger('UDSTool.' + __name__)

DEFAULT_HEADERS = (
    "Time",
    "Dir",
    "Type",
    "Destination IP",
    "Source IP",
    "DataLength",
    "Data",
    "ASCII"
)

DEFAULT_DISPLAY_HEADERS = (
    True,
    True,
    True,
    False,
    False,
    True,
    True,
    False
)


class DoIPTraceTableModel(QAbstractTableModel):
    """DoIP追踪表格模型，优化数据管理和批量更新"""

    def __init__(self, max_rows: int = 500):
        super().__init__()
        self.max_rows = max_rows  # 最大行数（防止内存溢出）
        self._data: List[DoIPMessageStruct] = []
        self._headers = DoIPMessageStruct().get_attr_names()[:8]

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:
        return len(self._data)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:
        return len(self._headers)

    def data(self, index: QModelIndex, role: int = Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None

        row = index.row()
        col = index.column()
        col_name = self._headers[col]

        # 显示数据
        if role == Qt.ItemDataRole.DisplayRole:
            if col_name == 'TX_id' or col_name == 'RX_id':
                return hex(getattr(self._data[row], col_name, '')).upper()[2:]
            else:
                return getattr(self._data[row], col_name, '')

    def append_trace_data(self, table_view_data: DoIPMessageStruct):
        """
        追加一行追踪数据，自动清理超出最大行数的旧数据
        :param table_view_data: 行数据
        """

        # 计算超出的行数，批量清理旧数据（支持一次性删除多行）
        excess_rows = len(self._data) - self.max_rows
        if excess_rows > 0:
            # 通知视图：即将删除从0到excess_rows-1的行（局部刷新）
            self.beginRemoveRows(QModelIndex(), 0, excess_rows - 1)
            # 切片批量删除旧行
            self._data = self._data[excess_rows:]
            # 通知视图：删除操作完成
            self.endRemoveRows()
            logger.debug(f"表格模型批量清理{excess_rows}行旧数据，当前剩余{len(self._data)}行")

        # 插入新行到模型末尾（局部刷新）
        insert_row_idx = len(self._data)  # 新行的索引（末尾）
        # 通知视图：即将在insert_row_idx位置插入1行
        self.beginInsertRows(QModelIndex(), insert_row_idx, insert_row_idx)
        self._data.append(table_view_data)
        # 通知视图：插入操作完成
        self.endInsertRows()
        logger.debug(f"表格模型新增1行数据，当前总行数{len(self._data)}")

    def headerData(self, section: int, orientation: Qt.Orientation,
                   role: int = Qt.ItemDataRole.DisplayRole):
        """重写表头方法，显示列名"""
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            if 0 <= section < len(self._headers):
                return self._headers[section]
        return None

    def clear(self):
        """清空表格数据（补全原有缺失的方法）"""
        self.beginResetModel()
        self._data.clear()
        self.endResetModel()
        logger.debug("表格模型数据已清空")

    @property
    def data_list(self) -> List[DoIPMessageStruct]:
        """获取原始数据（只读）"""
        return self._data.copy()


class DoIPTraceTableView(QTableView):
    """DoIP追踪表格，优化布局和交互"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._auto_scroll = True  # 自动滚动开关
        self._headers = DoIPMessageStruct().get_attr_names()[:8]
        self.trace_model: DoIPTraceTableModel = DoIPTraceTableModel(max_rows=500)  # 初始化模型
        # 列显示状态管理（key:列索引，value:是否显示）
        self._column_visible: Dict[int, bool] = {
            idx: True for idx in range(self.trace_model.columnCount())
        }

        self._header_column_mapping: Dict = {value: idx for idx, value in enumerate(self._headers)}

        self._init_ui()  # 初始化UI
        self._bind_scroll_listener()  # 绑定滚动监听
        self._init_header_context_menu()  # 初始化表头右键菜单
        self._init_data_context_menu()

    def _init_ui(self):
        """初始化表格UI属性，确保铺满布局"""
        # 选择行为
        self.setSelectionBehavior(self.SelectionBehavior.SelectRows)
        self.setAlternatingRowColors(True)  # 隔行变色

        # 模型设置
        self.setModel(self.trace_model)

        # 隐藏垂直表头（可选，节省空间）
        self.verticalHeader().setVisible(False)

        self.setWordWrap(False)  # 禁止自动换行
        self.setEditTriggers(self.EditTrigger.NoEditTriggers)  # 禁止编辑

        # 最后一列自适应剩余空间
        self.horizontalHeader().setStretchLastSection(True)

        try:
            for _id, state in enumerate(DEFAULT_DISPLAY_HEADERS):
                self._column_visible[_id] = state
                self.setColumnHidden(_id, not state)
            logger.debug(f"初始化列显示，{self._column_visible}")
        except Exception as e:
            logger.exception(f'初始化列显示失败：{str(e)}')

    def _bind_scroll_listener(self):
        """绑定滚动条监听，控制自动滚动"""
        scroll_bar: QScrollBar = self.verticalScrollBar()
        scroll_bar.valueChanged.connect(self._on_scroll_value_changed)

    # --------------------- 表头右键菜单（列选择） ---------------------
    def _init_header_context_menu(self):
        """初始化表头右键菜单：设置表头的上下文菜单策略，并绑定菜单事件"""
        # 设置水平表头的上下文菜单策略为自定义
        self.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        # 绑定表头右键菜单信号到槽函数
        self.horizontalHeader().customContextMenuRequested.connect(self._show_header_context_menu)

    def _show_header_context_menu(self, pos):
        """显示表头右键菜单：生成列的显示/隐藏选项"""
        menu = QMenu(self)  # 创建右键菜单

        # 为每列生成一个勾选的Action
        for col_idx, header_name in enumerate(self.trace_model._headers):
            action = QAction(header_name, self)
            action.setCheckable(True)  # 设置为可勾选
            action.setChecked(self._column_visible[col_idx])  # 根据当前状态设置勾选
            # 绑定Action的触发事件，传递列索引
            action.triggered.connect(lambda checked, idx=col_idx: self._toggle_column_visibility(idx, checked))
            menu.addAction(action)

        # 在鼠标右键位置显示菜单
        menu.exec(self.horizontalHeader().mapToGlobal(pos))

    def _toggle_column_visibility(self, col_idx: int, checked: bool):
        """切换列的显示/隐藏状态"""
        self._column_visible[col_idx] = checked
        # 调用QTableView的setColumnHidden接口控制列的可见性
        self.setColumnHidden(col_idx, not checked)
        logger.debug(f"列[{col_idx}:{self.trace_model._headers[col_idx]}]设置为{'显示' if checked else '隐藏'}")

    # --------------------- 数据区域右键菜单（复制/清空） ---------------------
    def _init_data_context_menu(self):
        """初始化数据区域右键菜单：复制、清空"""
        # 设置表格视图的上下文菜单策略为自定义（表头的策略已单独设置，不冲突）
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        # 绑定数据区域右键菜单信号
        self.customContextMenuRequested.connect(self._show_data_context_menu)

    def _show_data_context_menu(self, pos):
        """显示数据区域右键菜单"""
        # 获取右键位置对应的模型索引（判断是否点击在数据行上）
        index = self.indexAt(pos)
        menu = QMenu(self)

        # 若点击在数据行上，添加复制相关选项
        if index.isValid():
            # 复制单元格内容
            copy_cell_action = QAction("复制单元格内容", self)
            copy_cell_action.triggered.connect(lambda: self._copy_cell_data(index))
            menu.addAction(copy_cell_action)

            # 复制整行内容
            copy_row_action = QAction("复制整行内容", self)
            copy_row_action.triggered.connect(lambda: self._copy_row_data(index.row()))
            menu.addAction(copy_row_action)

            # 分隔线
            menu.addSeparator()

        export_text_action = QAction("导出到text", self)
        export_text_action.triggered.connect(lambda: self.export_to_text())
        menu.addAction(export_text_action)

        export_excel_action = QAction("导出到Excel", self)
        export_excel_action.triggered.connect(lambda: self.export_to_excel())
        menu.addAction(export_excel_action)

        # 清空数据选项（无论是否点击在数据行上都显示）
        clear_action = QAction("清空表格数据", self)
        clear_action.triggered.connect(self._clear_data)
        menu.addAction(clear_action)

        # 显示菜单（转换为全局坐标）
        menu.exec(self.mapToGlobal(pos))

    def export_to_text(self):
        """导出表格数据到Text文件，导出所有列"""
        # 若无数据则提示
        if not self.trace_model.data_list:
            QMessageBox.warning(self, "提示", "表格中无数据可导出！")
            return

        # 弹出文件保存对话框
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存Text文件",
            "DoIP_Trace.txt",
            "Text Files (*.txt);;All Files (*.*)"
        )
        if not file_path:
            return  # 用户取消保存

        # 获取所有列的索引和表头，不再过滤显示列
        all_cols = list(range(self.trace_model.columnCount()))  # 所有列的索引
        all_headers = [self.trace_model._headers[idx] for idx in all_cols]  # 所有列的显示表头

        # 写入文件
        try:
            with open(file_path, "w", encoding="utf-8") as f:
                # 写入表头（制表符分隔）
                f.write("\t".join(all_headers) + "\n")
                # 写入数据行（遍历所有列）
                for row_data in self.trace_model.data_list:
                    # 根据所有列的内部键名获取数据
                    all_row = [str(row_data.get(self.trace_model._headers[idx], "")) for idx in all_cols]
                    f.write("\t".join(all_row) + "\n")
            QMessageBox.information(self, "成功", f"数据已导出到：\n{file_path}")
            logger.debug(f"Text文件导出成功（所有列）：{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "失败", f"导出Text文件出错：{str(e)}")
            logger.error(f"Text文件导出失败：{str(e)}")

        # --------------------- 导出为Excel文件 ---------------------

    def export_to_excel(self):
        """导出表格数据到Excel文件，导出所有列"""
        # 若无数据则提示
        if not self.trace_model.data_list:
            QMessageBox.warning(self, "提示", "表格中无数据可导出！")
            return

        # 弹出文件保存对话框
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存Excel文件",
            "DoIP_Trace.xlsx",
            "Excel Files (*.xlsx);;All Files (*.*)"
        )
        if not file_path:
            return  # 用户取消保存

        # 获取所有列的索引和表头
        all_cols = list(range(self.trace_model.columnCount()))  # 所有列的索引
        all_headers = [self.trace_model._headers[idx] for idx in all_cols]  # 所有列的显示表头

        # 创建Excel工作簿并写入数据
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "DoIP Trace"

            # 写入表头并设置样式
            for col, header in enumerate(all_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                # 表头样式：加粗、居中、浅灰色背景
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = cell.fill.copy(fgColor="E0E0E0")  # 浅灰色

            # 写入数据行（遍历所有列）
            for row_idx, row_data in enumerate(self.trace_model.data_list, 2):
                # 【核心修改】根据所有列的内部键名获取数据
                all_row = [str(row_data.get(self.trace_model._headers[idx], "")) for idx in all_cols]
                for col_idx, value in enumerate(all_row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 限制最大列宽
                ws.column_dimensions[col_letter].width = adjusted_width

            # 保存文件
            wb.save(file_path)
            QMessageBox.information(self, "成功", f"数据已导出到：\n{file_path}")
            logger.debug(f"Excel文件导出成功（所有列）：{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "失败", f"导出Excel文件出错：{str(e)}")
            logger.error(f"Excel文件导出失败：{str(e)}")

    def _copy_cell_data(self, index: QModelIndex):
        """复制选中单元格的内容到剪贴板"""
        cell_data = index.data(Qt.ItemDataRole.DisplayRole) or ""
        clipboard = QApplication.clipboard()
        clipboard.setText(str(cell_data))
        logger.debug(f"复制单元格数据：{cell_data}")

    def _copy_row_data(self, row: int):
        """复制整行数据到剪贴板（按列名+值的格式）"""
        row_data = self.trace_model.data_list[row]
        headers = self.trace_model._headers
        # 拼接整行数据（仅包含显示的列）
        row_text = []
        for col_idx, (header, value) in enumerate(zip(headers, row_data)):
            if self._column_visible.get(col_idx, True):  # 只拼接显示的列
                row_text.append(f"{header}: {value}")
        row_str = "\n".join(row_text)

        clipboard = QApplication.clipboard()
        clipboard.setText(row_str)
        logger.debug(f"复制第{row}行数据：{row_str}")

    def _clear_data(self):
        """清空数据"""
        self.clear_trace_data()

    @Slot(int)
    def _on_scroll_value_changed(self, value: int):
        """根据滚动条位置更新自动滚动状态"""
        scroll_bar = self.verticalScrollBar()
        self._auto_scroll = (value == scroll_bar.maximum())
        if self._auto_scroll:
            logger.debug("表格滚动到底部，开启自动滚动")

    def add_trace_data(self, data: DoIPMessageStruct):
        """对外暴露的接口：添加追踪数据"""
        if data.is_empty():
            return

        try:
            if self._column_visible[self._header_column_mapping['ASCII']]:
                data.update_ascii_by_uds_data()
            self.trace_model.append_trace_data(data)
            if self._auto_scroll:
                self.scrollToBottom()  # 自动滚动到底部
        except Exception as e:
            logger.error(f"添加表格数据失败：{str(e)}")

    def clear_trace_data(self):
        """对外暴露的接口：清空表格数据"""
        self.trace_model.clear()

    @property
    def auto_scroll(self) -> bool:
        """获取自动滚动状态"""
        return self._auto_scroll

    @auto_scroll.setter
    def auto_scroll(self, state: bool):
        """设置自动滚动状态"""
        self._auto_scroll = state
        if state:
            self.scrollToBottom()
